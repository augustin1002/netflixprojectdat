"""
=============================================================
STEP 5: EXCEL DASHBOARD BUILDER
Netflix Data Engineering Project
=============================================================

CONCEPT:
--------
A dashboard is a visual summary of key metrics and trends.
We build it in Excel using openpyxl — Python's library for
reading and writing .xlsx files with full formatting control.

DASHBOARD COMPONENTS:
  📌 KPI Cards      → Single-number highlights (Total Titles, Movies, Shows)
  🥧 Pie Chart      → Movies vs TV Shows split
  📊 Bar Chart      → Top 10 Countries
  📈 Line Chart     → Content Added Per Year
  🎭 Genre Chart    → Top Genres horizontal bar

OPENPYXL KEY CONCEPTS:
  Workbook  → The whole Excel file
  Worksheet → A single tab/sheet in the file
  Cell      → A single cell (e.g., A1, B3)
  Chart     → Chart object embedded in a sheet
  Reference → Defines which cells to use for chart data
=============================================================
"""

import sqlite3
import pandas as pd
import numpy as np
import logging
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("output/pipeline.log", mode="a")
    ]
)
logger = logging.getLogger(__name__)

DB_PATH   = "output/netflix.db"
DASH_PATH = "dashboard/Netflix_Dashboard.xlsx"


# ── Color Palette (Netflix brand + professional) ──────────────────────────────
NETFLIX_RED    = "E50914"   # Netflix's iconic red
DARK_BG        = "141414"   # Netflix dark background
LIGHT_GRAY     = "F5F5F5"   # Card backgrounds
MED_GRAY       = "E0E0E0"   # Borders
WHITE          = "FFFFFF"
DARK_TEXT      = "1A1A1A"
ACCENT_BLUE    = "1565C0"
ACCENT_GREEN   = "2E7D32"
ACCENT_ORANGE  = "E65100"
HEADER_BG      = "B71C1C"   # Deep red for headers


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: style_cell()
# Makes applying multiple styles to one cell cleaner
# ══════════════════════════════════════════════════════════════════════════════
def style_cell(cell, bold=False, font_size=11, font_color="000000",
               bg_color=None, align="left", wrap=False, border=False):
    """Apply formatting styles to an openpyxl cell."""
    cell.font = Font(
        bold=bold,
        size=font_size,
        color=font_color,
        name="Arial"          # Professional consistent font
    )
    if bg_color:
        # PatternFill → solid color fill (like "Fill Color" button in Excel)
        cell.fill = PatternFill("solid", start_color=bg_color, end_color=bg_color)

    # Alignment → horizontal and vertical text positioning in cell
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=wrap
    )

    if border:
        thin = Side(style="thin", color=MED_GRAY)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: merge_and_style()
# Merge a range of cells and apply one style to the merged cell
# ══════════════════════════════════════════════════════════════════════════════
def merge_and_style(ws, cell_range, value="", bold=False, font_size=11,
                    font_color="000000", bg_color=None, align="center"):
    """Merge cells, write value, and apply styling."""
    ws.merge_cells(cell_range)
    # After merging, we only write to the TOP-LEFT cell of the range
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = value
    style_cell(top_left, bold=bold, font_size=font_size,
               font_color=font_color, bg_color=bg_color, align=align)


# ══════════════════════════════════════════════════════════════════════════════
# FUNCTION: fetch_data()
# Load all analysis data from SQLite for dashboard use
# ══════════════════════════════════════════════════════════════════════════════
def fetch_data(db_path: str) -> dict:
    """Fetch all required data from the database."""
    logger.info("Fetching data for dashboard...")
    conn = sqlite3.connect(db_path)

    data = {}
    try:
        data["kpis"] = pd.read_sql("""
            SELECT
                COUNT(*)                                             AS total_titles,
                SUM(CASE WHEN type='Movie'   THEN 1 ELSE 0 END)     AS total_movies,
                SUM(CASE WHEN type='TV Show' THEN 1 ELSE 0 END)     AS total_tv_shows,
                COUNT(DISTINCT country)                              AS unique_countries,
                COUNT(DISTINCT director)                             AS unique_directors,
                ROUND(AVG(content_age), 0)                          AS avg_content_age
            FROM netflix_titles
        """, conn)

        data["content_type"] = pd.read_sql("""
            SELECT type, COUNT(*) AS count
            FROM netflix_titles GROUP BY type ORDER BY count DESC
        """, conn)

        data["top_countries"] = pd.read_sql("""
            SELECT country, COUNT(*) AS total
            FROM netflix_titles
            WHERE country NOT IN ('Not Available','')
            GROUP BY country ORDER BY total DESC LIMIT 10
        """, conn)

        data["per_year"] = pd.read_sql("""
            SELECT added_year,
                   SUM(CASE WHEN type='Movie'   THEN 1 ELSE 0 END) AS movies,
                   SUM(CASE WHEN type='TV Show' THEN 1 ELSE 0 END) AS tv_shows,
                   COUNT(*) AS total
            FROM netflix_titles
            WHERE added_year >= 2008 AND added_year IS NOT NULL
            GROUP BY added_year ORDER BY added_year
        """, conn)

        data["ratings"] = pd.read_sql("""
            SELECT rating, COUNT(*) AS count FROM netflix_titles
            GROUP BY rating ORDER BY count DESC LIMIT 10
        """, conn)

        data["top_genres"] = pd.read_sql("""
            SELECT genre, count FROM summary_by_genre
            ORDER BY count DESC LIMIT 10
        """, conn)

        data["top_directors"] = pd.read_sql("""
            SELECT director, COUNT(*) AS titles
            FROM netflix_titles
            WHERE director NOT IN ('Not Available','')
            GROUP BY director ORDER BY titles DESC LIMIT 10
        """, conn)

    finally:
        conn.close()

    logger.info("Dashboard data fetched successfully.")
    return data


# ══════════════════════════════════════════════════════════════════════════════
# FUNCTION: build_title_sheet()
# Creates the Cover / Title page
# ══════════════════════════════════════════════════════════════════════════════
def build_title_sheet(wb: Workbook) -> None:
    """Build a professional title/cover sheet."""
    ws = wb.create_sheet("🎬 Cover", 0)
    ws.sheet_view.showGridLines = False  # Hide grid lines for clean look

    # Set column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in range(1, 30):
        ws.row_dimensions[row].height = 25

    # Dark background for entire sheet
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=11):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=DARK_BG, end_color=DARK_BG)

    # Netflix-style header bar
    for r in range(3, 7):
        for c in range(1, 12):
            ws.cell(r, c).fill = PatternFill("solid", start_color=NETFLIX_RED, end_color=NETFLIX_RED)

    # Title text
    merge_and_style(ws, "B4:J5", "🎬  NETFLIX DATA ENGINEERING PROJECT",
                    bold=True, font_size=24, font_color=WHITE, bg_color=NETFLIX_RED, align="center")

    merge_and_style(ws, "B6:J6", "End-to-End Data Pipeline  •  SQL Analysis  •  Excel Dashboard",
                    bold=False, font_size=13, font_color=WHITE, bg_color=NETFLIX_RED, align="center")

    # Description cards
    cards = [
        (9,  "B", "C", "📥  EXTRACT",    "Raw Netflix Excel dataset ingested with Pandas", ACCENT_BLUE),
        (13, "B", "C", "🔧  TRANSFORM",  "Cleaned, standardized & enriched with new columns", ACCENT_GREEN),
        (17, "B", "C", "🗄️  LOAD",       "ETL pipeline loads data into SQLite database", ACCENT_ORANGE),
        (21, "B", "C", "📊  ANALYZE",    "SQL queries answer 12 business questions", HEADER_BG),
    ]

    for row, c1, c2, title, desc, color in cards:
        ws.row_dimensions[row].height   = 30
        ws.row_dimensions[row+1].height = 25

        merge_and_style(ws, f"{c1}{row}:J{row}", title,
                        bold=True, font_size=14, font_color=WHITE, bg_color=color, align="left")
        merge_and_style(ws, f"{c1}{row+1}:J{row+1}", f"  {desc}",
                        bold=False, font_size=11, font_color=DARK_TEXT, bg_color=LIGHT_GRAY, align="left")

    # Footer
    merge_and_style(ws, "B27:J27",
                    "Built with Python  •  Pandas  •  SQLite  •  openpyxl  •  Data Engineering Project",
                    bold=False, font_size=10, font_color="888888", bg_color=DARK_BG, align="center")


# ══════════════════════════════════════════════════════════════════════════════
# FUNCTION: build_dashboard_sheet()
# Creates the main KPI + Charts dashboard
# ══════════════════════════════════════════════════════════════════════════════
def build_dashboard_sheet(wb: Workbook, data: dict) -> None:
    """Build the main KPI dashboard sheet."""
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [2, 18, 18, 18, 18, 18, 18, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── HEADER BAND ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 8

    for c in range(1, 9):
        ws.cell(2, c).fill = PatternFill("solid", start_color=NETFLIX_RED, end_color=NETFLIX_RED)

    merge_and_style(ws, "B2:G2", "🎬  NETFLIX CONTENT ANALYTICS DASHBOARD",
                    bold=True, font_size=20, font_color=WHITE, bg_color=NETFLIX_RED, align="center")

    # ── KPI CARDS (Row 4-7) ───────────────────────────────────────────────────
    kpi_row = data["kpis"].iloc[0]

    kpi_cards = [
        ("B", "Total Titles",    f"{int(kpi_row['total_titles']):,}",     HEADER_BG,      WHITE),
        ("C", "Total Movies",    f"{int(kpi_row['total_movies']):,}",     ACCENT_BLUE,    WHITE),
        ("D", "Total TV Shows",  f"{int(kpi_row['total_tv_shows']):,}",   ACCENT_GREEN,   WHITE),
        ("E", "Countries",       f"{int(kpi_row['unique_countries']):,}", ACCENT_ORANGE,  WHITE),
        ("F", "Directors",       f"{int(kpi_row['unique_directors']):,}", "6A1B9A",       WHITE),
        ("G", "Avg Content Age", f"{int(kpi_row['avg_content_age'])} yrs","37474F",       WHITE),
    ]

    # KPI card rows
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 15

    for col_letter, label, value, bg, fg in kpi_cards:
        # Label row
        cell_label = ws[f"{col_letter}5"]
        cell_label.value = label
        style_cell(cell_label, bold=False, font_size=10, font_color=fg, bg_color=bg, align="center")

        # Value row
        cell_value = ws[f"{col_letter}6"]
        cell_value.value = value
        style_cell(cell_value, bold=True, font_size=18, font_color=fg, bg_color=bg, align="center")

        # Padding rows
        for r in [4, 7]:
            ws[f"{col_letter}{r}"].fill = PatternFill("solid", start_color=bg, end_color=bg)

    # ── DATA TABLES FOR CHARTS ────────────────────────────────────────────────
    # We write data to hidden area of the sheet, then create charts from it.
    # Charts in openpyxl must reference cell ranges — they can't use Python objects directly.

    # Content Type data (col I, rows 10+)
    ws["I9"]  = "Content Type"; ws["J9"] = "Count"
    style_cell(ws["I9"], bold=True, bg_color=LIGHT_GRAY)
    style_cell(ws["J9"], bold=True, bg_color=LIGHT_GRAY)
    for i, row in data["content_type"].iterrows():
        ws[f"I{10+i}"] = row["type"];  ws[f"J{10+i}"] = int(row["count"])

    # Countries data (col I, rows 14+)
    ws["I14"] = "Country"; ws["J14"] = "Titles"
    style_cell(ws["I14"], bold=True, bg_color=LIGHT_GRAY)
    style_cell(ws["J14"], bold=True, bg_color=LIGHT_GRAY)
    for i, row in data["top_countries"].iterrows():
        ws[f"I{15+i}"] = row["country"]; ws[f"J{15+i}"] = int(row["total"])

    # Per-year data (col L)
    ws["L9"] = "Year"; ws["M9"] = "Movies"; ws["N9"] = "TV Shows"; ws["O9"] = "Total"
    for col in ["L9","M9","N9","O9"]:
        style_cell(ws[col], bold=True, bg_color=LIGHT_GRAY)
    for i, row in data["per_year"].iterrows():
        r = 10 + i
        ws[f"L{r}"] = int(row["added_year"])
        ws[f"M{r}"] = int(row["movies"])
        ws[f"N{r}"] = int(row["tv_shows"])
        ws[f"O{r}"] = int(row["total"])

    # Ratings data (col R)
    ws["R9"] = "Rating"; ws["S9"] = "Count"
    for col in ["R9","S9"]:
        style_cell(ws[col], bold=True, bg_color=LIGHT_GRAY)
    for i, row in data["ratings"].iterrows():
        ws[f"R{10+i}"] = row["rating"]; ws[f"S{10+i}"] = int(row["count"])

    nrows_type    = len(data["content_type"])
    nrows_country = len(data["top_countries"])
    nrows_year    = len(data["per_year"])
    nrows_ratings = len(data["ratings"])

    # ── PIE CHART: Movies vs TV Shows ─────────────────────────────────────────
    pie = PieChart()
    pie.title = "Movies vs TV Shows"
    pie.style = 10

    # Reference(ws, min_col, min_row, max_col, max_row)
    labels = Reference(ws, min_col=9, min_row=10, max_row=9+nrows_type)
    values = Reference(ws, min_col=10, min_row=9, max_row=9+nrows_type)
    pie.add_data(values, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 12; pie.width = 14
    ws.add_chart(pie, "B9")      # Anchor top-left corner of chart at B9

    # ── BAR CHART: Top Countries ──────────────────────────────────────────────
    bar = BarChart()
    bar.type    = "bar"           # Horizontal bars ("bar" = horizontal, "col" = vertical)
    bar.title   = "Top 10 Countries by Content"
    bar.y_axis.title = "Country"
    bar.x_axis.title = "Number of Titles"
    bar.style   = 10

    labels_c  = Reference(ws, min_col=9,  min_row=15, max_row=14+nrows_country)
    values_c  = Reference(ws, min_col=10, min_row=14, max_row=14+nrows_country)
    bar.add_data(values_c, titles_from_data=True)
    bar.set_categories(labels_c)
    bar.height = 14; bar.width = 20
    ws.add_chart(bar, "D9")

    # ── LINE CHART: Content Per Year ──────────────────────────────────────────
    line = LineChart()
    line.title        = "Netflix Content Growth by Year"
    line.y_axis.title = "Titles Added"
    line.x_axis.title = "Year"
    line.style        = 10

    years_ref   = Reference(ws, min_col=12, min_row=10, max_row=9+nrows_year)
    movies_ref  = Reference(ws, min_col=13, min_row=9,  max_row=9+nrows_year)
    shows_ref   = Reference(ws, min_col=14, min_row=9,  max_row=9+nrows_year)
    line.add_data(movies_ref, titles_from_data=True)
    line.add_data(shows_ref,  titles_from_data=True)
    line.set_categories(years_ref)
    line.height = 14; line.width = 22
    ws.add_chart(line, "B27")

    # ── BAR CHART: Ratings ────────────────────────────────────────────────────
    bar_r = BarChart()
    bar_r.type   = "col"          # Vertical column chart
    bar_r.title  = "Content Rating Distribution"
    bar_r.y_axis.title = "Count"
    bar_r.style  = 10

    labels_r = Reference(ws, min_col=18, min_row=10, max_row=9+nrows_ratings)
    values_r = Reference(ws, min_col=19, min_row=9,  max_row=9+nrows_ratings)
    bar_r.add_data(values_r, titles_from_data=True)
    bar_r.set_categories(labels_r)
    bar_r.height = 14; bar_r.width = 16
    ws.add_chart(bar_r, "F27")


# ══════════════════════════════════════════════════════════════════════════════
# FUNCTION: build_data_sheets()
# Write raw analysis tables to named sheets for drill-down
# ══════════════════════════════════════════════════════════════════════════════
def build_data_sheets(wb: Workbook, data: dict) -> None:
    """Write analysis data tables to individual sheets."""

    sheets = [
        ("🌍 Countries",     data["top_countries"]),
        ("📅 Year Analysis", data["per_year"]),
        ("⭐ Ratings",       data["ratings"]),
        ("🎥 Directors",     data["top_directors"]),
        ("🎭 Genres",        data["top_genres"]),
    ]

    for sheet_name, df in sheets:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        # Header row
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(1, col_idx, col_name.replace("_", " ").title())
            style_cell(cell, bold=True, font_size=11, font_color=WHITE,
                       bg_color=NETFLIX_RED, align="center", border=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        # Data rows (alternating colors for readability)
        for row_idx, row in df.iterrows():
            bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row_idx + 2, col_idx, value)
                style_cell(cell, font_size=10, bg_color=bg, border=True,
                           align="right" if isinstance(value, (int, float)) else "left")

        ws.row_dimensions[1].height = 28

    logger.info("Data sheets built.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    logger.info("===== STEP 5: EXCEL DASHBOARD BUILDER STARTED =====")

    os.makedirs("dashboard", exist_ok=True)

    # Fetch data
    data = fetch_data(DB_PATH)

    # Create workbook
    wb = Workbook()
    # Remove the default empty sheet Excel creates
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Build sheets
    build_title_sheet(wb, )
    build_dashboard_sheet(wb, data)
    build_data_sheets(wb, data)

    # Save the workbook
    wb.save(DASH_PATH)
    logger.info(f"Dashboard saved: {DASH_PATH}")

    print(f"\n✅ Excel Dashboard created: {DASH_PATH}")
    print("   Sheets created:")
    for s in wb.sheetnames:
        print(f"     • {s}")

    logger.info("===== STEP 5 COMPLETED SUCCESSFULLY =====\n")
